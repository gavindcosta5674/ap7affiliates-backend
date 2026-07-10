from fastapi import FastAPI, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta, date
from urllib.parse import quote
import csv
import uvicorn
import secrets
import io
import base64
import requests

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── FastAPI Setup ──────────────────────────────────────────────────────────
app = FastAPI(title="Master Affiliate Panel API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── In-Memory Databases ───────────────────────────────────────────────────
affiliates_db = []
affiliate_id_counter = 1000

# Initialize default affiliate
def init_default_affiliate():
    if not any(aff["id"] == "AFF-1001" for aff in affiliates_db):
        affiliates_db.append({
            "id": "AFF-1001",
            "affiliate_id": "AFF-1001",
            "first_name": "Marketing",
            "last_name": "Team",
            "username": "marketing",
            "password": "marketing123",
            "commission_pct": 0,
            "commission_percentage": 0,
            "status": "active",
            "total_players": 0,
            "total_deposits": 0,
            "total_withdrawals": 0,
            "total_bonuses": 0,
            "revenue": 0,
            "commission": 0,
            "created_at": "2025-06-07 12:00:00",
        })

init_default_affiliate()

# Links tracking
links_db = []
link_id_counter = 0

# Simple raw daily data sheet (admin can add rows, affiliate APIs can fetch them)
raw_data_db = []
raw_data_counter = 0

reports_db = {}

# Shared media gallery
media_db = []

# Players database
players_db = []
player_id_counter = 0

# ── Helper functions ──────────────────────────────────────────────────────
def find_affiliate(username: str, password: str):
    for aff in affiliates_db:
        if aff["username"] == username and aff["password"] == password and aff["status"] == "active":
            return aff
    return None


def find_affiliate_by_username(username: str):
    for aff in affiliates_db:
        if aff["username"].lower() == username.lower():
            return aff
    return None


def resolve_affiliate_identifier(value: str):
    if not value:
        return None
    text = value.strip()
    if not text:
        return None

    # Match raw affiliate id
    for aff in affiliates_db:
        if aff["id"].lower() == text.lower() or aff.get("affiliate_id", "").lower() == text.lower():
            return aff["id"]

    # Match username
    aff = find_affiliate_by_username(text)
    if aff:
        return aff["id"]

    # Match full name
    lower_text = text.lower()
    for aff in affiliates_db:
        full_name = f"{aff['first_name']} {aff['last_name']}".strip().lower()
        if full_name == lower_text:
            return aff["id"]

    return None


def normalize_csv_header(value: str) -> str:
    return value.strip().lower().replace('-', '_').replace(' ', '_')


def get_csv_value(cols: list[str], headers: Optional[list[str]], names: list[str], fallback_index: Optional[int] = None) -> str:
    if headers:
        for name in names:
            if name in headers:
                idx = headers.index(name)
                if idx < len(cols):
                    return cols[idx].strip()
    if fallback_index is not None and fallback_index < len(cols):
        return cols[fallback_index].strip()
    return ''


def parse_csv_float(value: str) -> float:
    try:
        return float(value) if value else 0.0
    except ValueError:
        return 0.0


def convert_xlsx_to_csv(file_bytes: bytes) -> str:
    """Convert XLSX file bytes to CSV string format"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=400, detail="XLSX support not available. Please upload CSV file or contact support.")
    
    try:
        workbook = load_workbook(io.BytesIO(file_bytes))
        worksheet = workbook.active
        
        csv_lines = []
        for row in worksheet.iter_rows(values_only=True):
            # Convert row values to strings, handling None values
            row_str = ','.join(str(cell) if cell is not None else '' for cell in row)
            csv_lines.append(row_str)
        
        return '\n'.join(csv_lines)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse XLSX file: {str(e)}")


def find_affiliate_by_id(aid: str):
    for aff in affiliates_db:
        if aff["id"] == aid:
            return aff
    return None


def resolve_report_dates(period: Optional[str], date_from: Optional[str], date_to: Optional[str]):
    if period and period != 'custom':
        today = date.today()
        if period == '7days':
            start = today - timedelta(days=6)
            end = today
        elif period == '15days':
            start = today - timedelta(days=14)
            end = today
        elif period == '30days':
            start = today - timedelta(days=29)
            end = today
        elif period == 'last_month':
            if today.month == 1:
                start = date(today.year - 1, 12, 1)
                end = date(today.year - 1, 12, 31)
            else:
                start = date(today.year, today.month - 1, 1)
                end = date(today.year, today.month, 1) - timedelta(days=1)
        elif period == '1year':
            start = today - timedelta(days=365)
            end = today
        else:
            start = None
            end = None
        return start.strftime('%Y-%m-%d') if start else None, end.strftime('%Y-%m-%d') if end else None

    return date_from, date_to


def build_report_entry(row: dict, affiliate_id: Optional[str] = None, commission_percentage: Optional[float] = None):
    deposit_count = int(row.get("count") or row.get("deposit_count") or 0)
    withdrawal_count = int(row.get("withdrawal_count") if row.get("withdrawal_count") is not None else (1 if float(row.get("withdrawal", 0) or 0) > 0 else 0))
    bonus_count = int(row.get("bonus_count") if row.get("bonus_count") is not None else (1 if float(row.get("bonus", 0) or 0) > 0 else 0))
    transaction_count = deposit_count + withdrawal_count + bonus_count
    entry = dict(row)
    if affiliate_id:
        entry["affiliate_id"] = affiliate_id
    entry["registration_date"] = entry.get("registration_date") or entry.get("registered_at") or entry.get("ftd_date") or ""
    entry["ftd_date"] = entry.get("ftd_date") or entry.get("first_deposit_date") or ""
    entry["first_deposit_amount"] = entry.get("first_deposit_amount") or 0
    entry["deposit_count"] = deposit_count
    entry["withdrawal_count"] = withdrawal_count
    entry["bonus_count"] = bonus_count
    entry["transaction_count"] = transaction_count
    if commission_percentage is not None:
        entry["commission_percentage"] = commission_percentage
        # Revenue = Deposits - Withdrawals - Bonuses
        revenue = round(float(entry.get("deposit", 0)) - float(entry.get("withdrawal", 0)) - float(entry.get("bonus", 0)), 2)
        entry["revenue"] = revenue
        entry["commission"] = round(revenue * float(commission_percentage) / 100, 2)
    return entry


def build_tracking_url(affiliate_id: str, platform: str, campaign: str) -> str:
    base = "http://ap7affiliates.online/"
    params = {
        "aff": affiliate_id,
        "source": platform,
        "campaign": campaign,
    }
    query = "&".join(f"{key}={quote(str(value), safe='')}" for key, value in params.items() if value)
    return f"{base}?{query}" if query else base

# ── Pydantic Models ───────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    success: bool
    token: str
    message: str
    affiliate: Optional[dict] = None

class CreateLinkRequest(BaseModel):
    name: str
    platform: str


class RawDataEntry(BaseModel):
    date: str
    affiliate_id: str
    player_username: str
    registration_date: Optional[str] = None
    deposit_amount: float = 0.0
    withdrawal_amount: float = 0.0
    bonus_amount: float = 0.0

class CsvImportRequest(BaseModel):
    csv_text: str
    affiliate_id: Optional[str] = None

class FileImportRequest(BaseModel):
    """Support both CSV text and base64-encoded XLSX files"""
    file_content: Optional[str] = None  # CSV text or base64 XLSX
    file_type: Optional[str] = "csv"  # 'csv' or 'xlsx'
    affiliate_id: Optional[str] = None

class PasswordUpdateRequest(BaseModel):
    affiliate_id: str
    password: str

class RegistrationImportRequest(BaseModel):
    csv_text: str
    affiliate_id: Optional[str] = None

class TransactionImportRequest(BaseModel):
    csv_text: str
    affiliate_id: Optional[str] = None


class RowsImportRequest(BaseModel):
    rows: List[dict]
    affiliate_id: Optional[str] = None
    import_type: Optional[str] = None

# =========================================================================
# MASTER ADMIN ENDPOINTS
# =========================================================================

@app.post("/api/login")
def master_login(body: LoginRequest):
    if body.username == "Chris9742" and body.password == "Chris9742":
        return LoginResponse(success=True, token="master-token-chris9742", message="Login successful")
    raise HTTPException(status_code=401, detail="Invalid username or password")

@app.get("/api/affiliates")
def get_affiliates():
    safe = []
    for aff in affiliates_db:
        aff_copy = {k: v for k, v in aff.items() if k != "password"}
        # Recalculate
        aff_copy["revenue"] = round(aff_copy.get("total_deposits", 0) - aff_copy.get("total_withdrawals", 0) - aff_copy.get("total_bonuses", 0), 2)
        aff_copy["commission"] = round(aff_copy["revenue"] * aff_copy.get("commission_pct", 20) / 100, 2)
        safe.append(aff_copy)
    return safe

@app.post("/api/affiliates")
def create_affiliate(body: dict):
    global affiliate_id_counter
    affiliate_id_counter += 1
    new_id = f"AFF-{affiliate_id_counter}"
    pct = float(body.get("percentage", 20))
    record = {
        "id": new_id,
        "affiliate_id": new_id,
        "first_name": body.get("first_name", ""),
        "last_name": body.get("last_name", ""),
        "username": body.get("username", ""),
        "password": body.get("password", ""),
        "commission_pct": pct,
        "commission_percentage": pct,
        "status": "active",
        "total_players": 0,
        "total_deposits": 0,
        "total_withdrawals": 0,
        "total_bonuses": 0,
        "revenue": 0,
        "commission": 0,
        "created_at": "2025-06-07 12:00:00",
    }
    affiliates_db.append(record)
    return {k: v for k, v in record.items() if k != "password"}

@app.put("/api/affiliates")
def update_affiliate_status(id: str, status: str):
    aff = find_affiliate_by_id(id)
    if not aff:
        raise HTTPException(status_code=404, detail="Affiliate not found")
    if status not in ["active", "inactive"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    aff["status"] = status
    return {"success": True, "message": f"Affiliate {id} status changed to {status}"}

@app.delete("/api/affiliates")
def delete_affiliate(id: str):
    global affiliates_db
    for i, aff in enumerate(affiliates_db):
        if aff["id"] == id:
            affiliates_db.pop(i)
            return {"success": True, "message": "Affiliate deleted"}
    raise HTTPException(status_code=404, detail="Affiliate not found")

@app.get("/api/dashboard")
def get_dashboard():
    total_affiliates = len(affiliates_db)
    active_affiliates = sum(1 for a in affiliates_db if a["status"] == "active")
    
    all_players = []
    unique_player_usernames = set()
    
    # Count from reports_db
    for aff_id, rows in reports_db.items():
        for r in rows:
            all_players.append(r)
            unique_player_usernames.add(r.get("player_username", "").lower())
    
    # Aggregate from raw_data_db (sum all transactions per player)
    player_aggregates = {}  # {player_username.lower(): {deposits, withdrawals, bonuses, ftd_date, reg_date, deposit_count, first_deposit_amount}}
    for row in raw_data_db:
        player_name = row.get("player_username", "").lower()
        if not player_name:
            continue
        
        if player_name not in player_aggregates:
            player_aggregates[player_name] = {
                "deposit": 0, 
                "withdrawal": 0, 
                "bonus": 0,
                "deposit_count": 0,
                "first_deposit_amount": 0,
                "ftd_date": None,
                "registration_date": row.get("registration_date", row.get("date", ""))
            }
        
        deposit = float(row.get("deposit_amount", 0) or 0)
        player_aggregates[player_name]["deposit"] += deposit
        player_aggregates[player_name]["withdrawal"] += float(row.get("withdrawal_amount", 0) or 0)
        player_aggregates[player_name]["bonus"] += float(row.get("bonus_amount", 0) or 0)
        
        # Track deposit count
        if deposit > 0:
            player_aggregates[player_name]["deposit_count"] += 1
            if not player_aggregates[player_name]["ftd_date"]:
                player_aggregates[player_name]["ftd_date"] = row.get("date", row.get("registration_date", ""))
                player_aggregates[player_name]["first_deposit_amount"] = deposit
    
    # Add aggregated raw_data players
    for player_name, totals in player_aggregates.items():
        if player_name not in unique_player_usernames:
            unique_player_usernames.add(player_name)
            all_players.append({
                "player_username": player_name,
                "deposit": totals["deposit"],
                "withdrawal": totals["withdrawal"],
                "bonus": totals["bonus"],
                "deposit_count": totals["deposit_count"],
                "ftd_date": totals["ftd_date"],
                "registration_date": totals["registration_date"],
                "first_deposit_amount": totals["first_deposit_amount"],
            })
    
    # Count from players_db
    for player in players_db:
        player_name = player.get("player_username", "").lower()
        if player_name and player_name not in unique_player_usernames:
            unique_player_usernames.add(player_name)
            all_players.append({
                "player_username": player.get("player_username", ""),
                "deposit": float(player.get("deposit_total", 0) or 0),
                "withdrawal": float(player.get("withdrawal_total", 0) or 0),
                "bonus": float(player.get("bonus_total", 0) or 0),
                "deposit_count": player.get("deposit_count", 0),
                "ftd_date": player.get("ftd_date", ""),
                "registration_date": player.get("registration_date", ""),
                "first_deposit_amount": player.get("first_deposit_amount", 0),
            })
    
    total_players = len(unique_player_usernames)
    total_deposits = sum(r.get("deposit", 0) for r in all_players)
    total_withdrawals = sum(r.get("withdrawal", 0) for r in all_players)
    total_bonuses = sum(r.get("bonus", 0) for r in all_players)
    total_revenue = round(total_deposits - total_withdrawals - total_bonuses, 2)
    
    avg_pct = sum(a["commission_pct"] for a in affiliates_db) / len(affiliates_db) if affiliates_db else 20
    total_commission = round(total_revenue * avg_pct / 100, 2)

    recent = []
    for aff in affiliates_db:
        aff_id = aff["id"]
        unique_players_aff = set()
        dep = 0
        wd = 0
        bn = 0
        
        # From reports_db
        rows = reports_db.get(aff_id, [])
        for r in rows:
            unique_players_aff.add(r.get("player_username", "").lower())
            dep += r.get("deposit", 0)
            wd += r.get("withdrawal", 0)
            bn += r.get("bonus", 0)
        
        # From raw_data_db - aggregate per player
        raw_players = {}
        for row in raw_data_db:
            if row.get("affiliate_id") == aff_id:
                player_name = row.get("player_username", "").lower()
                if player_name not in raw_players:
                    raw_players[player_name] = {"deposit": 0, "withdrawal": 0, "bonus": 0}
                raw_players[player_name]["deposit"] += float(row.get("deposit_amount", 0) or 0)
                raw_players[player_name]["withdrawal"] += float(row.get("withdrawal_amount", 0) or 0)
                raw_players[player_name]["bonus"] += float(row.get("bonus_amount", 0) or 0)
        
        for player_name, totals in raw_players.items():
            unique_players_aff.add(player_name)
            dep += totals["deposit"]
            wd += totals["withdrawal"]
            bn += totals["bonus"]
        
        # From players_db
        for player in players_db:
            if player.get("affiliate_id") == aff_id:
                player_name = player.get("player_username", "").lower()
                unique_players_aff.add(player_name)
                dep += float(player.get("deposit_total", 0) or 0)
                wd += float(player.get("withdrawal_total", 0) or 0)
                bn += float(player.get("bonus_total", 0) or 0)
        
        rev = round(dep - wd - bn, 2)
        comm = round(rev * aff["commission_pct"] / 100, 2)
        recent.append({
            "affiliate_id": aff_id,
            "first_name": aff["first_name"],
            "last_name": aff["last_name"],
            "total_players": len(unique_players_aff),
            "total_deposits": round(dep, 2),
            "total_withdrawals": round(wd, 2),
            "total_bonuses": round(bn, 2),
            "revenue": rev,
            "commission": comm,
            "commission_percentage": aff["commission_pct"],
        })

    return {
        "total_affiliates": total_affiliates,
        "active_affiliates": active_affiliates,
        "total_players": total_players,
        "total_deposits": round(total_deposits, 2),
        "total_withdrawals": round(total_withdrawals, 2),
        "total_bonuses": round(total_bonuses, 2),
        "total_revenue": total_revenue,
        "total_commission": total_commission,
        "recent_activity": recent,
    }

@app.get("/api/reports")
def get_reports(affiliate_id: str = None, date_from: str = None, date_to: str = None, player_username: str = None, period: str = None, authorization: Optional[str] = Header(None, alias="Authorization")):
    # Check if admin requesting all data
    token = authorization
    if authorization and authorization.lower().startswith("bearer "):
        token = authorization[7:]
    
    # If admin token, allow viewing all affiliates
    if is_admin_token(token):
        affiliate_id = None  # Show all
    
    date_from, date_to = resolve_report_dates(period, date_from, date_to)
    all_reports = []
    seen_players = set()  # Track unique player_username
    
    # PRIORITY 1: Aggregate transactions from raw_data_db by player (most accurate data)
    player_aggregates = {}  # {(player_username, affiliate_id): {deposits, withdrawals, bonuses, dates, deposit_count}}
    for row in raw_data_db:
        player_name = row.get("player_username", "").lower()
        aff_id = row.get("affiliate_id", "")
        key = (player_name, aff_id)
        
        if key not in player_aggregates:
            player_aggregates[key] = {
                "deposits": 0.0,
                "withdrawals": 0.0,
                "bonuses": 0.0,
                "ftd_date": None,
                "first_deposit_amount": 0.0,
                "registration_date": row.get("registration_date", row.get("date", "")),
                "player_username": row.get("player_username", ""),
                "affiliate_id": aff_id,
                "deposit_count": 0,
            }
        
        deposit = float(row.get("deposit_amount", 0) or 0)
        withdrawal = float(row.get("withdrawal_amount", 0) or 0)
        bonus = float(row.get("bonus_amount", 0) or 0)
        
        # Aggregate amounts
        player_aggregates[key]["deposits"] += deposit
        player_aggregates[key]["withdrawals"] += withdrawal
        player_aggregates[key]["bonuses"] += bonus
        
        # Track FTD date and first deposit amount
        if deposit > 0:
            player_aggregates[key]["deposit_count"] += 1
            if not player_aggregates[key]["ftd_date"]:
                player_aggregates[key]["ftd_date"] = row.get("date", row.get("registration_date", ""))
                player_aggregates[key]["first_deposit_amount"] = deposit
    
    # Add aggregated players from raw_data_db to reports
    for idx, ((player_name, aff_id), data) in enumerate(player_aggregates.items(), start=1):
        if affiliate_id and aff_id != affiliate_id:
            continue
        if player_name in seen_players:
            continue
        seen_players.add(player_name)
        
        aff = find_affiliate_by_id(aff_id)
        commission_percentage = aff["commission_pct"] if aff else 20
        
        deposit_total = round(data["deposits"], 2)
        withdrawal_total = round(data["withdrawals"], 2)
        bonus_total = round(data["bonuses"], 2)
        revenue = round(deposit_total - withdrawal_total - bonus_total, 2)
        commission = round(revenue * commission_percentage / 100, 2)
        
        all_reports.append(build_report_entry({
            "id": f"RAW-{idx}",
            "player_id": f"RAW-{idx}",
            "player_username": data["player_username"],
            "affiliate_id": aff_id,
            "ftd_date": data["ftd_date"] or data["registration_date"] or "",
            "registration_date": data["registration_date"] or "",
            "first_deposit_amount": data["first_deposit_amount"],
            "deposit": deposit_total,
            "count": data["deposit_count"],
            "deposit_count": data["deposit_count"],
            "withdrawal": withdrawal_total,
            "withdrawal_count": 1 if withdrawal_total > 0 else 0,
            "bonus": bonus_total,
            "bonus_count": 1 if bonus_total > 0 else 0,
            "revenue": revenue,
            "commission": commission,
        }, affiliate_id=aff_id, commission_percentage=commission_percentage))

    # PRIORITY 2: Add players from reports_db
    for aff_id, rows in reports_db.items():
        if affiliate_id and aff_id != affiliate_id:
            continue
        for r in rows:
            player_name = r.get("player_username", "").lower()
            if player_name in seen_players:
                continue
            seen_players.add(player_name)
            
            aff = find_affiliate_by_id(aff_id)
            commission_percentage = aff["commission_pct"] if aff else 20
            all_reports.append(build_report_entry(r, affiliate_id=aff_id, commission_percentage=commission_percentage))

    # PRIORITY 3: Add players from players_db (only if not already added)
    for player in players_db:
        aff_id = player.get("affiliate_id")
        if affiliate_id and aff_id != affiliate_id:
            continue
        player_name = player.get("player_username", "").lower()
        if player_name in seen_players:
            continue
        seen_players.add(player_name)
        
        aff = find_affiliate_by_id(aff_id)
        commission_percentage = aff["commission_pct"] if aff else 20
        row = {
            "id": player.get("player_id"),
            "player_id": player.get("player_id"),
            "player_username": player.get("player_username"),
            "affiliate_id": aff_id,
            "ftd_date": player.get("ftd_date") or "",
            "registration_date": player.get("registration_date") or player.get("ftd_date") or "",
            "first_deposit_amount": player.get("first_deposit_amount", 0),
            "deposit": player.get("deposit_total", 0),
            "count": player.get("deposit_count", 0),
            "deposit_count": player.get("deposit_count", 0),
            "withdrawal": player.get("withdrawal_total", 0),
            "withdrawal_count": 1 if float(player.get("withdrawal_total", 0) or 0) > 0 else 0,
            "bonus": player.get("bonus_total", 0),
            "bonus_count": 1 if float(player.get("bonus_total", 0) or 0) > 0 else 0,
            "revenue": round(float(player.get("deposit_total", 0) or 0) - float(player.get("withdrawal_total", 0) or 0) - float(player.get("bonus_total", 0) or 0), 2),
            "commission": player.get("commission", 0),
        }
        all_reports.append(build_report_entry(row, affiliate_id=aff_id, commission_percentage=commission_percentage))

    if player_username:
        all_reports = [r for r in all_reports if player_username.lower() in r["player_username"].lower()]
    if date_from:
        all_reports = [r for r in all_reports if r.get("ftd_date", "") >= date_from]
    if date_to:
        all_reports = [r for r in all_reports if r.get("ftd_date", "") <= date_to]

    return all_reports

@app.get("/api/commission")
def get_commission():
    total_deposits = 0
    total_withdrawals = 0
    total_bonuses = 0
    for aff_id, rows in reports_db.items():
        for r in rows:
            total_deposits += r["deposit"]
            total_withdrawals += r["withdrawal"]
            total_bonuses += r["bonus"]
    total_revenue = round(total_deposits - total_withdrawals - total_bonuses, 2)
    avg_pct = sum(a["commission_pct"] for a in affiliates_db) / len(affiliates_db) if affiliates_db else 20
    total_commission = round(total_revenue * avg_pct / 100, 2)

    aff_data = []
    for aff in affiliates_db:
        aff_id = aff["id"]
        rows = reports_db.get(aff_id, [])
        dep = sum(r["deposit"] for r in rows)
        wd = sum(r["withdrawal"] for r in rows)
        bn = sum(r["bonus"] for r in rows)
        rev = round(dep - wd - bn, 2)
        comm = round(rev * aff["commission_pct"] / 100, 2)
        aff_data.append({
            "affiliate_id": aff_id,
            "first_name": aff["first_name"],
            "last_name": aff["last_name"],
            "total_players": len(rows),
            "total_deposits": dep,
            "total_withdrawals": wd,
            "total_bonuses": bn,
            "revenue": rev,
            "commission_percentage": aff["commission_pct"],
            "commission": comm,
        })

    return {
        "total_revenue": total_revenue,
        "total_commission": total_commission,
        "average_percentage": round(avg_pct, 2),
        "affiliates": aff_data,
    }

@app.post("/api/commission")
def update_commission(body: dict):
    global affiliates_db
    if "global_percentage" in body:
        pct = float(body["global_percentage"])
        for aff in affiliates_db:
            aff["commission_pct"] = pct
        return {"success": True, "message": "Global commission updated", "percentage": pct}
    if "affiliate_id" in body and "percentage" in body:
        aff = find_affiliate_by_id(body["affiliate_id"])
        if not aff:
            raise HTTPException(status_code=404, detail="Affiliate not found")
        aff["commission_pct"] = float(body["percentage"])
        return {"success": True, "message": f"Commission updated for {body['affiliate_id']}"}
    raise HTTPException(status_code=400, detail="Invalid request")

@app.get("/api/media")
def get_media():
    return media_db

@app.post("/api/media")
def upload_media():
    return {"error": "File upload not supported via in-memory API", "media_db": media_db}

@app.delete("/api/media")
def delete_media(id: int):
    global media_db
    for i, m in enumerate(media_db):
        if m["id"] == id:
            media_db.pop(i)
            return {"success": True, "message": "Media deleted"}
    raise HTTPException(status_code=404, detail="Media not found")

@app.get("/api/players")
def get_players(affiliate_id: str = None):
    """Get all players with aggregated transaction data"""
    all_players = []
    seen_players = set()
    
    # PRIORITY 1: Aggregate transactions from raw_data_db by player (most accurate data)
    player_aggregates = {}  # {(player_username, affiliate_id): {...}}
    for row in raw_data_db:
        player_name = row.get("player_username", "").lower()
        aff_id = row.get("affiliate_id", "")
        key = (player_name, aff_id)
        
        if key not in player_aggregates:
            player_aggregates[key] = {
                "deposits": 0.0,
                "withdrawals": 0.0,
                "bonuses": 0.0,
                "ftd_date": None,
                "first_deposit_amount": 0.0,
                "last_deposit_date": None,
                "registration_date": row.get("registration_date", row.get("date", "")),
                "player_username": row.get("player_username", ""),
                "affiliate_id": aff_id,
                "deposit_count": 0,
            }
        
        deposit = float(row.get("deposit_amount", 0) or 0)
        withdrawal = float(row.get("withdrawal_amount", 0) or 0)
        bonus = float(row.get("bonus_amount", 0) or 0)
        
        # Aggregate amounts
        player_aggregates[key]["deposits"] += deposit
        player_aggregates[key]["withdrawals"] += withdrawal
        player_aggregates[key]["bonuses"] += bonus
        
        # Track FTD date and first deposit amount
        if deposit > 0:
            player_aggregates[key]["deposit_count"] += 1
            if not player_aggregates[key]["ftd_date"]:
                player_aggregates[key]["ftd_date"] = row.get("date", row.get("registration_date", ""))
                player_aggregates[key]["first_deposit_amount"] = deposit
            # Always update last_deposit_date
            player_aggregates[key]["last_deposit_date"] = row.get("date", row.get("registration_date", ""))
    
    # Add aggregated players from raw_data_db
    for idx, ((player_name, aff_id), data) in enumerate(player_aggregates.items(), start=1):
        if affiliate_id and aff_id != affiliate_id:
            continue
        if player_name in seen_players:
            continue
        seen_players.add(player_name)
        
        aff = find_affiliate_by_id(aff_id)
        commission_percentage = aff["commission_pct"] if aff else 20
        
        deposit_total = round(data["deposits"], 2)
        withdrawal_total = round(data["withdrawals"], 2)
        bonus_total = round(data["bonuses"], 2)
        revenue = round(deposit_total - withdrawal_total - bonus_total, 2)
        commission = round(revenue * commission_percentage / 100, 2)
        
        all_players.append({
            "id": f"RAW-{idx}",
            "player_id": f"RAW-{idx}",
            "player_username": data["player_username"],
            "affiliate_id": aff_id,
            "registration_date": data["registration_date"] or "",
            "ftd_date": data["ftd_date"] or data["registration_date"] or "",
            "first_deposit_amount": data["first_deposit_amount"],
            "last_deposit_date": data["last_deposit_date"] or data["ftd_date"] or "",
            "deposit": round(data["deposits"], 2),
            "deposit_count": data["deposit_count"],
            "withdrawal": round(data["withdrawals"], 2),
            "bonus": round(data["bonuses"], 2),
            "revenue": round(round(data["deposits"], 2) - round(data["withdrawals"], 2) - round(data["bonuses"], 2), 2),
            "commission": commission,
            "commission_percentage": commission_percentage,
        })
    
    # Add players from players_db (only if not already added)
    for player in players_db:
        aff_id = player.get("affiliate_id")
        if affiliate_id and aff_id != affiliate_id:
            continue
        player_name = player.get("player_username", "").lower()
        if player_name in seen_players:
            continue
        seen_players.add(player_name)
        
        aff = find_affiliate_by_id(aff_id)
        commission_percentage = aff["commission_pct"] if aff else 20
        
        deposit_total = float(player.get("deposit_total", 0) or 0)
        withdrawal_total = float(player.get("withdrawal_total", 0) or 0)
        bonus_total = float(player.get("bonus_total", 0) or 0)
        revenue = round(deposit_total - withdrawal_total - bonus_total, 2)
        commission = round(revenue * commission_percentage / 100, 2)
        
        all_players.append({
            "id": player.get("player_id"),
            "player_id": player.get("player_id"),
            "player_username": player.get("player_username"),
            "affiliate_id": aff_id,
            "registration_date": player.get("registration_date", ""),
            "ftd_date": player.get("ftd_date", ""),
            "first_deposit_amount": player.get("first_deposit_amount", 0),
            "last_deposit_date": player.get("last_deposit_date", player.get("ftd_date", "")),
            "deposit": deposit_total,
            "deposit_count": player.get("deposit_count", 0),
            "withdrawal": withdrawal_total,
            "bonus": bonus_total,
            "revenue": revenue,
            "commission": commission,
            "commission_percentage": commission_percentage,
        })
    
    return all_players

@app.post("/api/players")
def create_player(body: dict):
    global player_id_counter, players_db
    player_id_counter += 1
    player_id = f"PLR{player_id_counter}"
    aff_id = body.get("affiliate_id", "")
    
    aff = find_affiliate_by_id(aff_id)
    if not aff:
        raise HTTPException(status_code=404, detail="Affiliate not found")
    
    deposit = float(body.get("deposit_amount", 0))
    withdrawal = float(body.get("withdrawal_amount", 0))
    bonus = float(body.get("bonus_amount", 0))
    revenue = round(deposit - withdrawal - bonus, 2)
    commission = round(revenue * aff["commission_pct"] / 100, 2)
    
    record = {
        "player_id": player_id,
        "player_username": body.get("player_username", ""),
        "affiliate_id": aff_id,
        "ftd_date": body.get("ftd_date", "2025-06-07"),
        "registration_date": body.get("registration_date", body.get("ftd_date", "2025-06-07")),
        "deposit_total": deposit,
        "deposit_count": 1 if deposit > 0 else 0,
        "withdrawal_total": withdrawal,
        "bonus_total": bonus,
        "revenue": revenue,
        "commission": commission,
    }
    players_db.append(record)
    
    # Also add to reports_db
    if aff_id not in reports_db:
        reports_db[aff_id] = []
    reports_db[aff_id].append({
        "id": player_id_counter + 100,
        "affiliate_id": aff_id,
        "player_username": body.get("player_username", ""),
        "ftd_date": body.get("ftd_date", "2025-06-07"),
        "registration_date": body.get("registration_date", body.get("ftd_date", "2025-06-07")),
        "deposit": deposit,
        "count": 1 if deposit > 0 else 0,
        "withdrawal": withdrawal,
        "bonus": bonus,
        "revenue": revenue,
        "commission": commission,
        "commission_percentage": aff["commission_pct"],
    })
    
    return {"success": True, "player_id": player_id, **record}

@app.delete("/api/players")
def delete_player(id: str):
    global players_db
    for i, p in enumerate(players_db):
        if p["player_id"] == id:
            players_db.pop(i)
            return {"success": True, "message": "Player deleted"}
    raise HTTPException(status_code=404, detail="Player not found")

# =========================================================================
# AFFILIATE PARTNER ENDPOINTS
# =========================================================================

SESSION_TOKENS = {}

def is_admin_token(token: Optional[str]) -> bool:
    """Check if token is for master admin (Chris9742)"""
    return token == "master-token-chris9742"

def get_affiliate_id_from_header(authorization: Optional[str] = Header(None, alias="Authorization")) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    token = authorization
    if authorization.lower().startswith("bearer "):
        token = authorization[7:]
    affiliate_id = SESSION_TOKENS.get(token)
    if not affiliate_id:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return affiliate_id

@app.post("/api/affiliate/login")
def affiliate_login(body: LoginRequest):
    aff = find_affiliate_by_username(body.username)
    if not aff:
        raise HTTPException(status_code=401, detail="Incorrect Username")
    if aff["password"] != body.password:
        raise HTTPException(status_code=401, detail="Incorrect Password")
    if aff["status"] != "active":
        raise HTTPException(status_code=401, detail="Account is inactive")
    token = secrets.token_hex(16)
    SESSION_TOKENS[token] = aff["id"]
    safe = {k: v for k, v in aff.items() if k != "password"}
    return {"success": True, "token": token, "message": "Affiliate login successful", "affiliate": safe}

@app.get("/api/affiliate/dashboard")
def affiliate_dashboard(authorization: Optional[str] = Header(None, alias="Authorization")):
    affiliate_id = get_affiliate_id_from_header(authorization)
    aff = find_affiliate_by_id(affiliate_id)
    if not aff:
        raise HTTPException(status_code=404, detail="Affiliate not found")
    rows = reports_db.get(affiliate_id, [])
    total_deposits = sum(r["deposit"] for r in rows)
    total_withdrawals = sum(r["withdrawal"] for r in rows)
    total_bonuses = sum(r["bonus"] for r in rows)
    total_revenue = total_deposits - total_withdrawals - total_bonuses
    total_commission = sum(r["commission"] for r in rows)
    return {
        "affiliate_id": affiliate_id,
        "first_name": aff["first_name"],
        "last_name": aff["last_name"],
        "commission_pct": aff["commission_pct"],
        "total_players": len(rows),
        "total_deposits": round(total_deposits, 2),
        "total_withdrawals": round(total_withdrawals, 2),
        "total_bonuses": round(total_bonuses, 2),
        "total_revenue": round(total_revenue, 2),
        "total_commission": round(total_commission, 2),
        "status": aff["status"],
    }

@app.get("/api/affiliate/reports")
def affiliate_reports(authorization: Optional[str] = Header(None, alias="Authorization"), date_from: str = None, date_to: str = None, period: str = None):
    affiliate_id = get_affiliate_id_from_header(authorization)
    date_from, date_to = resolve_report_dates(period, date_from, date_to)
    result = []
    seen_players = set()  # Track unique player_username
    aff = find_affiliate_by_id(affiliate_id)
    commission_percentage = aff["commission_pct"] if aff else 20
    
    # PRIORITY 1: Process raw_data_db first (has actual transaction data)
    for row in raw_data_db:
        if row.get("affiliate_id") != affiliate_id:
            continue
        player_name = row.get("player_username", "").lower()
        if player_name in seen_players:
            continue
        seen_players.add(player_name)
        
        deposit = float(row.get("deposit_amount", 0) or 0)
        withdrawal = float(row.get("withdrawal_amount", 0) or 0)
        bonus = float(row.get("bonus_amount", 0) or 0)
        revenue = round(deposit - withdrawal - bonus, 2)
        result.append(build_report_entry({
            "id": f"RAW-{len(result)}",
            "player_id": f"RAW-{len(result)}",
            "player_username": row.get("player_username", ""),
            "affiliate_id": row.get("affiliate_id", ""),
            "ftd_date": row.get("date") or row.get("registration_date") or "",
            "registration_date": row.get("registration_date") or row.get("date") or "",
            "deposit": deposit,
            "count": 1 if deposit > 0 else 0,
            "withdrawal": withdrawal,
            "withdrawal_count": 1 if withdrawal > 0 else 0,
            "bonus": bonus,
            "bonus_count": 1 if bonus > 0 else 0,
            "revenue": revenue,
            "commission": round(revenue * commission_percentage / 100, 2),
        }, affiliate_id=affiliate_id, commission_percentage=commission_percentage))

    if date_from:
        result = [r for r in result if r.get("ftd_date", "") >= date_from]
    if date_to:
        result = [r for r in result if r.get("ftd_date", "") <= date_to]
    return result

@app.get("/api/raw-data")
def get_raw_data():
    return raw_data_db

@app.delete("/api/admin/clear-all-data")
def admin_clear_all_data(authorization: Optional[str] = Header(None, alias="Authorization")):
    """Admin only: Clear all imported data (keeps default affiliate)"""
    token = authorization
    if authorization and authorization.lower().startswith("bearer "):
        token = authorization[7:]
    
    if not is_admin_token(token):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    global raw_data_db, players_db, reports_db
    
    raw_data_db.clear()
    players_db.clear()
    reports_db.clear()
    
    return {
        "success": True, 
        "message": "All data cleared successfully",
        "note": "Default affiliate (AFF-1001) remains intact"
    }


@app.post("/api/raw-data/import-csv")
def import_raw_data_from_csv(body: CsvImportRequest):
    if not body.csv_text or not body.csv_text.strip():
        raise HTTPException(status_code=400, detail="CSV text is required")

    rows = []
    errors = []
    reader = csv.reader(body.csv_text.splitlines())
    col_indices = {}
    first_row = True
    
    for line_num, raw_cols in enumerate(reader, start=1):
        cols = [c.strip() for c in raw_cols]
        if not any(cols):
            continue

        # Parse headers - build flexible column mapping
        if first_row:
            normalized = [normalize_csv_header(c) for c in cols]
            # Build flexible column index mapping
            for idx, norm_header in enumerate(normalized):
                if norm_header in ['datetime', 'date_time', 'date', 'timestamp']:
                    col_indices['date'] = idx
                elif norm_header in ['deposit', 'deposit_amount', 'credit']:
                    col_indices['deposit'] = idx
                elif norm_header in ['withdraw', 'withdrawal', 'withdrawal_amount', 'debit']:
                    col_indices['withdraw'] = idx
                elif norm_header in ['bonus', 'bonus_amount']:
                    col_indices['bonus'] = idx
                elif norm_header in ['to', 'player', 'player_username', 'username']:
                    col_indices['username'] = idx
                elif norm_header in ['affiliate', 'affiliate_id', 'affiliate_username', 'aff']:
                    col_indices['affiliate'] = idx
            first_row = False
            continue
        
        if is_transaction := any(k in col_indices for k in ['deposit', 'date']):
            # Transaction format: Extract using flexible mapping
            date_value = cols[col_indices.get('date', 0)].strip() if 'date' in col_indices else (cols[0].strip() if cols else '')
            deposit = parse_csv_float(cols[col_indices.get('deposit', 4)].strip() if 'deposit' in col_indices else (cols[4] if len(cols) > 4 else '0'))
            withdraw = parse_csv_float(cols[col_indices.get('withdraw', 5)].strip() if 'withdraw' in col_indices else (cols[5] if len(cols) > 5 else '0'))
            bonus = parse_csv_float(cols[col_indices.get('bonus', 6)].strip() if 'bonus' in col_indices else (cols[6] if len(cols) > 6 else '0'))
            player_username = cols[col_indices.get('username', 2)].strip() if 'username' in col_indices else (cols[2].strip() if len(cols) > 2 else '')
            
            # Get affiliate: from request body, or from mapped column, or from column 1
            affiliate_id = body.affiliate_id
            if not affiliate_id and 'affiliate' in col_indices:
                affiliate_candidate = cols[col_indices.get('affiliate')].strip()
                if affiliate_candidate:
                    affiliate_id = affiliate_candidate
        
        if not player_username:
            continue

        # Default to AFF-1001 (Marketing) if no affiliate specified
        if not affiliate_id:
            affiliate_id = "AFF-1001"

        rows.append({
            'date': date_value,
            'affiliate_id': affiliate_id,
            'player_username': player_username,
            'registration_date': date_value,
            'deposit_amount': round(deposit, 2),
            'withdrawal_amount': round(withdraw, 2),
            'bonus_amount': round(bonus, 2),
        })

    if errors:
        error_msg = "❌ " + " | ".join(errors)
        return {"success": False, "error": error_msg}

    for row in rows:
        create_raw_data_record(row)

    return {'success': True, 'imported_count': len(rows), 'rows': rows}

@app.put("/api/affiliates/password")
def update_affiliate_password(body: PasswordUpdateRequest):
    aff = find_affiliate_by_id(body.affiliate_id)
    if not aff:
        raise HTTPException(status_code=404, detail="Affiliate not found")
    aff["password"] = body.password
    return {"success": True, "message": f"Password updated for {body.affiliate_id}"}

@app.post("/api/raw-data/import-registrations")
def import_registration_data_from_csv(body: RegistrationImportRequest):
    if not body.csv_text or not body.csv_text.strip():
        raise HTTPException(status_code=400, detail="CSV text is required")

    rows = []
    errors = []
    duplicates = []
    reader = csv.reader(body.csv_text.splitlines())
    headers = None
    first_row = True
    
    # Get existing player usernames
    existing_players = set(row.get('player_username', '').lower() for row in raw_data_db if row.get('player_username'))

    for line_num, raw_cols in enumerate(reader, start=1):
        cols = [c.strip() for c in raw_cols]
        if not any(cols):
            continue
        
        if first_row:
            normalized = [normalize_csv_header(c) for c in cols]
            header_names = {
                'sno', 'clientid', 'phoneno', 'signupdate', 'affiliate_id', 'affiliate_username',
                'date', 'datetime', 'timestamp', 'player', 'player_username', 'username', 'to',
                'affiliate', 'aff'
            }
            if any(name in normalized for name in header_names):
                headers = normalized
                first_row = False
                continue
            first_row = False

        # Map columns: Sno, ClientId, PhoneNo, SignupDate, Affiliate ID, Affiliate Username
        # Or: DateTime, Deposit, Withdraw, Bonus, remark, To, Affiliate
        sno = get_csv_value(cols, headers, ['sno', 'serial', 'id'], 0)
        player_username = get_csv_value(cols, headers, ['clientid', 'client_id', 'player_username', 'username', 'to'], 1)
        phone_or_deposit = get_csv_value(cols, headers, ['phoneno', 'phone', 'deposit'], 2)
        signup_or_withdraw = get_csv_value(cols, headers, ['signupdate', 'signup_date', 'date', 'withdraw'], 3)
        affiliate_id_or_bonus = get_csv_value(cols, headers, ['affiliate_id', 'aff_id', 'bonus'], 4)
        affiliate_username = get_csv_value(cols, headers, ['affiliate_username', 'affiliate', 'aff'], 5)
        
        if not player_username:
            continue

        # Check for duplicate player
        if player_username.lower() in existing_players:
            duplicates.append(f"Line {line_num}: Player '{player_username}' already registered. Skipped.")
            continue

        # Determine which format: Registration or Transaction
        is_registration = headers and any(name in headers for name in ['sno', 'clientid', 'phoneno'])
        
        if is_registration:
            # Registration format
            date_value = signup_or_withdraw  # SignupDate column
            affiliate_id = affiliate_id_or_bonus  # Affiliate ID column
        else:
            # Fallback for transaction-like format
            date_value = signup_or_withdraw
            affiliate_id = affiliate_id_or_bonus if affiliate_id_or_bonus else affiliate_username

        # Use request body affiliate_id if provided
        if body.affiliate_id:
            affiliate_id = body.affiliate_id
        
        # Default to AFF-1001 (Marketing) if no affiliate specified
        if not affiliate_id:
            affiliate_id = "AFF-1001"

        if player_username.lower() in existing_players:
            duplicates.append(f"Line {line_num}: Player '{player_username}' already registered. Skipped")
            continue

        rows.append({
            'date': date_value,
            'affiliate_id': affiliate_id,
            'player_username': player_username,
            'registration_date': date_value,
            'deposit_amount': 0.0,
            'withdrawal_amount': 0.0,
            'bonus_amount': 0.0,
        })
        existing_players.add(player_username.lower())

    if errors:
        error_msg = "❌ " + " | ".join(errors)
        if duplicates:
            error_msg += " ⚠️ " + " | ".join(duplicates)
        return {"success": False, "error": error_msg}

    for row in rows:
        create_raw_data_record(row)

    result = {'success': True, 'imported_count': len(rows), 'rows': rows}
    if duplicates:
        result['warnings'] = duplicates
    return result

@app.post("/api/raw-data/import-registrations-force")
def import_registrations_force(body: RegistrationImportRequest):
    """Force import registrations, ignoring duplicates"""
    if not body.csv_text or not body.csv_text.strip():
        raise HTTPException(status_code=400, detail="CSV text is required")

    rows = []
    reader = csv.reader(body.csv_text.splitlines())
    headers = None
    first_row = True
    
    for line_num, raw_cols in enumerate(reader, start=1):
        cols = [c.strip() for c in raw_cols]
        if not any(cols):
            continue
        
        if first_row:
            normalized = [normalize_csv_header(c) for c in cols]
            headers = normalized
            first_row = False
            continue

        # Extract values - Registration format
        date_value = get_csv_value(cols, headers, ['signupdate', 'signup_date', 'date'], 3)
        player_username = get_csv_value(cols, headers, ['clientid', 'client_id', 'username'], 1)
        
        if not player_username:
            continue

        # Get affiliate ID
        affiliate_id = body.affiliate_id or (cols[4].strip() if len(cols) > 4 else None)
        if not affiliate_id:
            affiliate_id = "AFF-1001"

        rows.append({
            'date': date_value,
            'affiliate_id': affiliate_id,
            'player_username': player_username,
            'registration_date': date_value,
            'deposit_amount': 0.0,
            'withdrawal_amount': 0.0,
            'bonus_amount': 0.0,
        })

    # Import all rows (no duplicate check)
    for row in rows:
        create_raw_data_record(row)

    return {'success': True, 'imported_count': len(rows), 'rows': rows, 'message': f'Imported {len(rows)} registrations (force mode - duplicates ignored)'}

@app.post("/api/raw-data/import-transactions")
def import_transaction_data_from_csv(body: TransactionImportRequest):
    if not body.csv_text or not body.csv_text.strip():
        raise HTTPException(status_code=400, detail="CSV text is required")

    rows = []
    errors = []
    existing_dates = set()
    reader = csv.reader(body.csv_text.splitlines())
    headers = None
    first_row = True
    col_indices = {}  # Map column names to indices
    
    # Get all existing transaction dates
    for entry in raw_data_db:
        if entry.get('date'):
            existing_dates.add(entry['date'])

    for line_num, raw_cols in enumerate(reader, start=1):
        cols = [c.strip() for c in raw_cols]
        if not any(cols):
            continue

        if first_row:
            # Build column index map by searching for header names
            normalized_headers = [normalize_csv_header(c) for c in cols]
            
            # Find column indices by header name
            for idx, norm_header in enumerate(normalized_headers):
                if norm_header in ['datetime', 'date_time', 'date', 'timestamp']:
                    col_indices['date'] = idx
                elif norm_header in ['deposit', 'deposit_amount', 'credit']:
                    col_indices['deposit'] = idx
                elif norm_header in ['withdraw', 'withdrawal', 'withdrawal_amount', 'debit']:
                    col_indices['withdraw'] = idx
                elif norm_header in ['bonus', 'bonus_amount']:
                    col_indices['bonus'] = idx
                elif norm_header in ['username', 'player_username', 'player', 'to']:
                    col_indices['username'] = idx
                elif norm_header in ['affiliate', 'affiliate_id', 'affiliate_username', 'aff']:
                    col_indices['affiliate'] = idx
                elif norm_header in ['remark', 'note', 'remarks']:
                    col_indices['remark'] = idx
            
            headers = normalized_headers
            first_row = False
            continue

        # Extract values using column indices (flexible order)
        date_value = cols[col_indices['date']].strip() if 'date' in col_indices else cols[0].strip()
        deposit = parse_csv_float(cols[col_indices['deposit']].strip() if 'deposit' in col_indices else (cols[1] if len(cols) > 1 else '0'))
        withdraw = parse_csv_float(cols[col_indices['withdraw']].strip() if 'withdraw' in col_indices else (cols[2] if len(cols) > 2 else '0'))
        bonus = parse_csv_float(cols[col_indices['bonus']].strip() if 'bonus' in col_indices else (cols[3] if len(cols) > 3 else '0'))
        player_username = cols[col_indices['username']].strip() if 'username' in col_indices else (cols[4].strip() if len(cols) > 4 else '')
        
        if not player_username:
            continue

        # Get affiliate: from request body, or from Affiliate column, or default to AFF-1001
        affiliate_id = body.affiliate_id
        if not affiliate_id and 'affiliate' in col_indices:
            affiliate_candidate = cols[col_indices['affiliate']].strip()
            if affiliate_candidate:
                affiliate_id = affiliate_candidate
        
        # Default to AFF-1001 (Marketing) if no affiliate specified
        if not affiliate_id:
            affiliate_id = "AFF-1001"
        
        rows.append({
            'date': date_value,
            'affiliate_id': affiliate_id,
            'player_username': player_username,
            'registration_date': date_value,
            'deposit_amount': round(deposit, 2),
            'withdrawal_amount': round(withdraw, 2),
            'bonus_amount': round(bonus, 2),
        })

    # Check for existing dates and inform user
    dates_in_import = set(row['date'] for row in rows if row['date'])
    overlapping_dates = dates_in_import & existing_dates
    
    info_msg = ""
    if overlapping_dates:
        sorted_dates = sorted(list(overlapping_dates))
        info_msg = f"⚠️ Existing transaction data found for dates: {', '.join(sorted_dates)}. Proceeding with import will update these entries."

    # Auto-create player registrations for new players
    for row in rows:
        player_username = row['player_username']
        affiliate_id = row['affiliate_id']
        registration_date = row['registration_date']
        
        # Check if player already exists in players_db
        existing_player = next((p for p in players_db if p.get('player_username', '').lower() == player_username.lower() and p.get('affiliate_id') == affiliate_id), None)
        if not existing_player:
            # Create new player registration
            players_db.append({
                'player_id': f"P-{len(players_db) + 1}",
                'player_username': player_username,
                'affiliate_id': affiliate_id,
                'registration_date': registration_date,
                'ftd_date': registration_date,
                'deposit_total': 0.0,
                'deposit_count': 0,
                'withdrawal_total': 0.0,
                'bonus_total': 0.0,
                'revenue': 0.0,
                'commission': 0.0,
            })
        
        create_raw_data_record(row)

    result = {'success': True, 'imported_count': len(rows), 'rows': rows}
    if info_msg:
        result['info'] = info_msg
    return result


@app.post("/api/raw-data/import-file")
def import_file_universal(body: FileImportRequest):
    """Universal import endpoint that handles both CSV and XLSX files"""
    if not body.file_content or not body.file_content.strip():
        raise HTTPException(status_code=400, detail="file_content is required")

    # Convert file content to CSV text
    csv_text = body.file_content
    
    if body.file_type and body.file_type.lower() == 'xlsx':
        try:
            import base64
            file_bytes = base64.b64decode(body.file_content)
            csv_text = convert_xlsx_to_csv(file_bytes)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to process XLSX file: {str(e)}")
    
    # Process as CSV
    rows = []
    errors = []
    existing_dates = set()
    reader = csv.reader(csv_text.splitlines())
    headers = None
    first_row = True
    
    # Get all existing transaction dates
    for entry in raw_data_db:
        if entry.get('date'):
            existing_dates.add(entry['date'])

    for line_num, raw_cols in enumerate(reader, start=1):
        cols = [c.strip() for c in raw_cols]
        if not any(cols):
            continue

        if first_row:
            normalized = [normalize_csv_header(c) for c in cols]
            header_names = {'date', 'datetime', 'timestamp', 'deposit', 'withdraw', 'withdrawal', 'bonus', 'player', 'to', 'affiliate', 'affiliate_id', 'marketing'}
            if any(name in normalized for name in header_names):
                headers = normalized
                first_row = False
                continue
            first_row = False

        date_value = get_csv_value(cols, headers, ['datetime', 'date_time', 'date', 'timestamp'], 0)
        deposit = parse_csv_float(get_csv_value(cols, headers, ['deposit', 'deposit_amount', 'credit'], 1))
        withdraw = parse_csv_float(get_csv_value(cols, headers, ['withdraw', 'withdrawal', 'withdrawal_amount', 'debit'], 2))
        bonus = parse_csv_float(get_csv_value(cols, headers, ['bonus', 'bonus_amount'], 3))
        player_username = get_csv_value(cols, headers, ['to', 'player', 'player_username', 'username'], 4)
        if not player_username:
            continue

        # Get affiliate_id: from request body first, then CSV
        affiliate_id = body.affiliate_id
        if not affiliate_id:
            candidate = get_csv_value(cols, headers, ['affiliate', 'affiliate_id', 'aff', 'marketing', 'source'], len(cols) - 1 if len(cols) > 5 else None)
            if candidate:
                affiliate_id = candidate
        
        if not affiliate_id:
            errors.append(f"Line {line_num}: No affiliate ID found.")
            continue
        
        resolved_id = resolve_affiliate_identifier(affiliate_id)
        if not resolved_id:
            errors.append(f"Line {line_num}: Affiliate '{affiliate_id}' not found.")
            continue
        affiliate_id = resolved_id

        rows.append({
            'date': date_value,
            'affiliate_id': affiliate_id,
            'player_username': player_username,
            'registration_date': date_value,
            'deposit_amount': round(deposit, 2),
            'withdrawal_amount': round(withdraw, 2),
            'bonus_amount': round(bonus, 2),
        })

    if errors:
        error_msg = "❌ " + " | ".join(errors)
        return {"success": False, "error": error_msg}

    # Check for existing dates
    dates_in_import = set(row['date'] for row in rows if row['date'])
    overlapping_dates = dates_in_import & existing_dates
    
    info_msg = ""
    if overlapping_dates:
        sorted_dates = sorted(list(overlapping_dates))
        info_msg = f"⚠️ Existing transaction data found for dates: {', '.join(sorted_dates)}. Proceeding with import will update these entries."

    # Auto-create player registrations
    for row in rows:
        player_username = row['player_username']
        affiliate_id = row['affiliate_id']
        registration_date = row['registration_date']
        
        existing_player = next((p for p in players_db if p.get('player_username', '').lower() == player_username.lower() and p.get('affiliate_id') == affiliate_id), None)
        if not existing_player:
            players_db.append({
                'player_id': f"P-{len(players_db) + 1}",
                'player_username': player_username,
                'affiliate_id': affiliate_id,
                'registration_date': registration_date,
                'ftd_date': registration_date,
                'deposit_total': 0.0,
                'deposit_count': 0,
                'withdrawal_total': 0.0,
                'bonus_total': 0.0,
                'revenue': 0.0,
                'commission': 0.0,
            })
        
        create_raw_data_record(row)

    result = {'success': True, 'imported_count': len(rows), 'rows': rows}
    if info_msg:
        result['info'] = info_msg
    return result


@app.post("/api/raw-data/import-rows")
def import_rows(body: RowsImportRequest):
    if not body.rows or len(body.rows) == 0:
        raise HTTPException(status_code=400, detail="No rows provided")

    created = []
    for r in body.rows:
        try:
            date_value = r.get('date') or r.get('registration_date') or ''
            player_username = r.get('player_username') or r.get('player') or r.get('username') or ''
            deposit = float(r.get('deposit_amount') or r.get('deposit') or 0) if r.get('deposit_amount') is not None else 0.0
            withdrawal = float(r.get('withdrawal_amount') or r.get('withdrawal') or 0) if r.get('withdrawal_amount') is not None else 0.0
            bonus = float(r.get('bonus_amount') or r.get('bonus') or 0) if r.get('bonus_amount') is not None else 0.0

            if body.affiliate_id:
                affiliate_id = body.affiliate_id
            else:
                candidate = r.get('affiliate') or r.get('affiliate_id') or r.get('marketing') or ''
                affiliate_id = resolve_affiliate_identifier(candidate) or 'AFF-1004'

            entry = {
                'date': date_value,
                'affiliate_id': affiliate_id,
                'player_username': player_username,
                'registration_date': date_value,
                'deposit_amount': round(deposit, 2),
                'withdrawal_amount': round(withdrawal, 2),
                'bonus_amount': round(bonus, 2),
            }
            rec = create_raw_data_record(entry)
            created.append(rec)
        except Exception:
            continue

    return {'success': True, 'imported_count': len(created), 'rows': created}


def create_raw_data_record(entry: dict):
    global raw_data_counter
    raw_data_counter += 1
    record = {
        'id': raw_data_counter,
        'date': entry['date'],
        'affiliate_id': entry['affiliate_id'],
        'player_username': entry['player_username'],
        'registration_date': entry.get('registration_date') or entry['date'],
        'deposit_amount': round(float(entry.get('deposit_amount', 0) or 0), 2),
        'withdrawal_amount': round(float(entry.get('withdrawal_amount', 0) or 0), 2),
        'bonus_amount': round(float(entry.get('bonus_amount', 0) or 0), 2),
    }
    raw_data_db.append(record)
    return record


@app.post("/api/raw-data")
def create_raw_data(entry: RawDataEntry):
    return create_raw_data_record({
        'date': entry.date,
        'affiliate_id': entry.affiliate_id,
        'player_username': entry.player_username,
        'registration_date': entry.registration_date,
        'deposit_amount': entry.deposit_amount,
        'withdrawal_amount': entry.withdrawal_amount,
        'bonus_amount': entry.bonus_amount,
    })


@app.get("/api/affiliate/raw-data")
def affiliate_get_raw_data(authorization: Optional[str] = Header(None, alias="Authorization")):
    affiliate_id = get_affiliate_id_from_header(authorization)
    return [row for row in raw_data_db if row.get("affiliate_id") == affiliate_id]


@app.get("/api/affiliate/media")
def affiliate_get_media(authorization: Optional[str] = Header(None, alias="Authorization")):
    get_affiliate_id_from_header(authorization)
    return media_db

@app.get("/api/links")
def get_links():
    return links_db

@app.post("/api/links")
def create_link(body: CreateLinkRequest, authorization: Optional[str] = Header(None, alias="Authorization")):
    global link_id_counter, links_db
    affiliate_id = get_affiliate_id_from_header(authorization)
    link_id_counter += 1
    safe_name = body.name.lower().replace(" ", "_")
    safe_platform = body.platform.lower().replace(" ", "_")
    tracking_url = build_tracking_url(affiliate_id, safe_platform, safe_name)
    record = {
        "id": link_id_counter,
        "name": body.name,
        "platform": body.platform,
        "affiliate_id": affiliate_id,
        "tracking_url": tracking_url,
        "created_at": "2025-06-07 12:00:00",
    }
    links_db.append(record)
    return record

@app.delete("/api/links/{link_id}")
def delete_link(link_id: int, authorization: Optional[str] = Header(None, alias="Authorization")):
    affiliate_id = get_affiliate_id_from_header(authorization)
    global links_db
    for i, link in enumerate(links_db):
        if link["id"] == link_id:
            if link["affiliate_id"] != affiliate_id:
                raise HTTPException(status_code=403, detail="Not authorized to delete this link")
            links_db.pop(i)
            return {"success": True, "message": "Link deleted"}
    raise HTTPException(status_code=404, detail="Link not found")

@app.get("/api/affiliate/links")
def get_affiliate_links(authorization: Optional[str] = Header(None, alias="Authorization")):
    affiliate_id = get_affiliate_id_from_header(authorization)
    return [l for l in links_db if l["affiliate_id"] == affiliate_id]

# ── Google Sheets Integration ──────────────────────────────────────────────
def convert_google_sheets_url_to_csv_export(sheet_url: str) -> str:
    """Convert Google Sheets URL to CSV export URL"""
    import re
    # Extract sheet ID from URL
    # Format: https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit?gid={GID}
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheet_url)
    if not match:
        raise ValueError("Invalid Google Sheets URL format")
    
    sheet_id = match.group(1)
    # Extract gid (sheet number) if present, default to 0
    # Look for ?gid=, #gid=, or &gid=
    gid_match = re.search(r'[?#&]gid=([0-9]+)', sheet_url)
    gid = gid_match.group(1) if gid_match else '0'
    
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

@app.post("/api/raw-data/import-from-google-sheets")
def import_from_google_sheets(body: dict):
    """
    Import data from Google Sheets (registrations or transactions)
    
    Request body:
    {
        "sheet_url": "https://docs.google.com/spreadsheets/d/YOUR_SHEET_ID/edit#gid=0",
        "data_type": "registration" or "transaction",
        "affiliate_id": "AFF-1001" (optional)
    }
    """
    sheet_url = body.get("sheet_url")
    data_type = body.get("data_type", "registration")  # registration or transaction
    affiliate_id = body.get("affiliate_id")
    
    if not sheet_url:
        raise HTTPException(status_code=400, detail="sheet_url is required")
    
    try:
        # Convert to CSV export URL and fetch
        csv_url = convert_google_sheets_url_to_csv_export(sheet_url)
        response = __import__('requests').get(csv_url, timeout=10)
        response.raise_for_status()
        csv_text = response.text
        
        if not csv_text.strip():
            raise HTTPException(status_code=400, detail="Google Sheet is empty")
        
        # Route to appropriate importer
        if data_type == "transaction":
            # Import as transactions
            payload = {"csv_text": csv_text, "affiliate_id": affiliate_id}
            return import_transaction_data_from_csv(TransactionImportRequest(**payload))
        else:
            # Import as registrations (default)
            payload = {"csv_text": csv_text, "affiliate_id": affiliate_id}
            return import_registrations_force(RegistrationImportRequest(**payload))
            
    except Exception as e:
        error_msg = str(e)
        if "Invalid Google Sheets URL" in error_msg:
            raise HTTPException(status_code=400, detail=error_msg)
        raise HTTPException(status_code=500, detail=f"Failed to fetch from Google Sheets: {error_msg}")

# ── Entry Point ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)