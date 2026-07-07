import io
import csv
import sys
from pathlib import Path
import openpyxl

# Ensure project root is on sys.path so `app` can be imported when running this script
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app import app
from fastapi.testclient import TestClient

TRANSACTION_XLSX = ROOT / 'Transaction Affiliate Data.xlsx'
REGISTRATION_XLSX = ROOT / 'Registration Affiliate Data.xlsx'

client = TestClient(app)


def xlsx_to_csv_text(path: Path) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    out = io.StringIO()
    writer = csv.writer(out)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(["" if v is None else str(v) for v in row])
    return out.getvalue()


def print_preview(csv_text: str, lines=8):
    print('--- Preview (first lines) ---')
    for i, line in enumerate(csv_text.splitlines()):
        if i >= lines:
            break
        print(line)
    print('-----------------------------\n')


def test_import_transactions():
    if not TRANSACTION_XLSX.exists():
        print('Transaction file not found:', TRANSACTION_XLSX)
        return
    csv_text = xlsx_to_csv_text(TRANSACTION_XLSX)
    print('Transaction file -> CSV conversion done.')
    print_preview(csv_text, lines=6)
    resp = client.post('/api/raw-data/import-transactions', json={'csv_text': csv_text, 'affiliate_id': ''})
    print('Status:', resp.status_code)
    data = resp.json()
    print('Imported count:', data.get('imported_count'))
    rows = data.get('rows', [])
    print('First imported rows:')
    for r in rows[:10]:
        print('-', r)
    print('\n')


def test_import_registrations():
    if not REGISTRATION_XLSX.exists():
        print('Registration file not found:', REGISTRATION_XLSX)
        return
    csv_text = xlsx_to_csv_text(REGISTRATION_XLSX)
    print('Registration file -> CSV conversion done.')
    print_preview(csv_text, lines=6)
    resp = client.post('/api/raw-data/import-registrations', json={'csv_text': csv_text, 'affiliate_id': ''})
    print('Status:', resp.status_code)
    data = resp.json()
    print('Imported count:', data.get('imported_count'))
    rows = data.get('rows', [])
    print('First imported rows:')
    for r in rows[:10]:
        print('-', r)
    print('\n')


if __name__ == '__main__':
    print('\nRunning import tests against uploaded Excel files...\n')
    test_import_transactions()
    test_import_registrations()
    print('Done.')
